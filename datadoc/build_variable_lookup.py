"""
Build a unified variable lookup from SSB variabellister (Excel) and metadata_1191.csv.

Produces:
  - variabel_beskrivelser.csv: all variables from SSB Excel files with descriptions
  - variabel_lookup.csv: metadata_1191 variables enriched with SSB descriptions
"""

import openpyxl
import pandas as pd
import os
import re
import csv

XLDIR = "variabellister_ssb"
METADATA = "metadata_1191.csv"

# ── Step 1: Parse all Excel files ──────────────────────────────────────────────

def parse_variabelliste(filepath):
    """Extract (variabelnavn, beskrivelse, definisjon, kodeliste) from an SSB Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    source = os.path.basename(filepath)

    # Find the sheet with variables
    target_sheet = None
    for sn in wb.sheetnames:
        if "velg" in sn.lower() or "Velg" in sn:
            target_sheet = sn
            break
    if target_sheet is None:
        wb.close()
        return []

    ws = wb[target_sheet]
    variables = []

    for row in ws.iter_rows(values_only=True):
        # Row must have content in column A (beskrivelse) and column D (variabelnavn)
        if len(row) < 4:
            continue
        beskrivelse = row[0]
        variabelnavn = row[3]

        if not variabelnavn or not beskrivelse:
            continue

        beskrivelse = str(beskrivelse).strip()
        variabelnavn = str(variabelnavn).strip()

        # Skip header rows and section headers
        if variabelnavn in ("Variabelnavn", "Intern informasjon", ""):
            continue
        if beskrivelse in ("Variabelbeskrivelse", ""):
            continue
        # Skip rows where variabelnavn looks like a note/comment (long text, spaces)
        if len(variabelnavn) > 60 or variabelnavn.startswith("*"):
            continue

        # Get definition and kodeliste if available
        definisjon = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        kodeliste = str(row[7]).strip() if len(row) > 7 and row[7] else ""

        # Get available years if present
        fra_aar = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        til_aar = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        # Clean up "definisjon" links (just mark as having definition)
        if definisjon.lower() == "definisjon":
            definisjon = "ja"
        if kodeliste.lower() in ("kodeliste", "koder"):
            kodeliste = "ja"

        variables.append({
            "variabelnavn_ssb": variabelnavn,
            "beskrivelse": beskrivelse,
            "definisjon": definisjon,
            "kodeliste": kodeliste,
            "fra_aar": fra_aar,
            "til_aar": til_aar,
            "kilde_fil": source,
        })

    wb.close()
    return variables


print("Parsing SSB Excel files...")
all_vars = []
for f in sorted(os.listdir(XLDIR)):
    if not f.endswith((".xlsx", ".xlsm")):
        continue
    path = os.path.join(XLDIR, f)
    try:
        variables = parse_variabelliste(path)
        print(f"  {f}: {len(variables)} variables")
        all_vars.extend(variables)
    except Exception as e:
        print(f"  {f}: ERROR - {e}")

ssb_df = pd.DataFrame(all_vars)
print(f"\nTotal SSB variables parsed: {len(ssb_df)}")
print(f"Unique variable names: {ssb_df.variabelnavn_ssb.nunique()}")

# ── Step 2: Normalize variable names for matching ──────────────────────────────

def normalize_varname(name):
    """
    Normalize SSB variable names so they can match metadata_1191 varnames.
    SSB uses patterns like 'wxx_xxxx_lopenr_person' or 'wyrkinnt'
    while metadata has 'lopenr_person' or 'wyrkinnt'.
    """
    name = name.lower().strip()
    # Remove wxx_xxxx_ prefix (placeholder for year)
    name = re.sub(r'^wxx_xxxx_', '', name)
    # Remove w prefix for single-word vars like 'wyrkinnt' -> keep as is
    return name

ssb_df["varname_normalized"] = ssb_df.variabelnavn_ssb.apply(normalize_varname)

# Save the full SSB variable descriptions
ssb_df.to_csv("variabel_beskrivelser.csv", index=False, encoding="utf-8-sig")
print(f"\nSaved: variabel_beskrivelser.csv ({len(ssb_df)} rows)")

# ── Step 3: Load metadata and build lookup ─────────────────────────────────────

print(f"\nLoading {METADATA}...")
meta_df = pd.read_csv(METADATA, encoding="latin-1", on_bad_lines="skip")
print(f"  {len(meta_df)} rows, {meta_df.varname.nunique()} unique variable names")

# Build a deduplicated SSB description lookup (keep first/best match per normalized name)
# Prefer entries with actual descriptions over section headers
ssb_lookup = (
    ssb_df[ssb_df.beskrivelse.str.len() > 2]
    .drop_duplicates(subset="varname_normalized", keep="first")
    .set_index("varname_normalized")
)

# Match metadata variables to SSB descriptions
meta_df["varname_lower"] = meta_df.varname.str.lower().str.strip()
meta_df = meta_df.merge(
    ssb_lookup[["beskrivelse", "kilde_fil", "definisjon", "kodeliste"]],
    left_on="varname_lower",
    right_index=True,
    how="left",
)
meta_df.rename(columns={"beskrivelse": "ssb_beskrivelse", "kilde_fil": "ssb_kilde"}, inplace=True)

matched = meta_df.ssb_beskrivelse.notna().sum()
total = len(meta_df)
unique_matched = meta_df.loc[meta_df.ssb_beskrivelse.notna(), "varname"].nunique()
unique_total = meta_df.varname.nunique()
print(f"\nMatched: {matched}/{total} rows ({matched/total*100:.1f}%)")
print(f"Unique variables matched: {unique_matched}/{unique_total} ({unique_matched/unique_total*100:.1f}%)")

# ── Step 4: Build compact lookup for quick reading ─────────────────────────────

# One row per unique variable, combining info from both sources
compact = (
    meta_df.groupby("varname")
    .agg(
        varlabel=("varlabel", "first"),
        ssb_beskrivelse=("ssb_beskrivelse", "first"),
        ssb_kilde=("ssb_kilde", "first"),
        vartype=("vartype", "first"),
        n_files=("filename", "nunique"),
        example_file=("filename", "first"),
        definisjon=("definisjon", "first"),
        kodeliste=("kodeliste", "first"),
    )
    .reset_index()
)

# Use SSB description if available, otherwise fall back to Stata varlabel
compact["description"] = compact.ssb_beskrivelse.fillna(compact.varlabel)

# Reorder columns for readability
compact = compact[
    ["varname", "description", "varlabel", "ssb_beskrivelse", "vartype",
     "n_files", "example_file", "ssb_kilde", "definisjon", "kodeliste"]
].sort_values("varname")

compact.to_csv("variabel_lookup.csv", index=False, encoding="utf-8-sig")
print(f"\nSaved: variabel_lookup.csv ({len(compact)} unique variables)")

# ── Step 5: Summary stats ─────────────────────────────────────────────────────

has_desc = compact.description.notna().sum()
print(f"\nVariables with description: {has_desc}/{len(compact)} ({has_desc/len(compact)*100:.1f}%)")
print(f"Variables with SSB description: {compact.ssb_beskrivelse.notna().sum()}")
print(f"Variables with Stata varlabel only: {(compact.ssb_beskrivelse.isna() & compact.varlabel.notna()).sum()}")
print(f"Variables with no description: {compact.description.isna().sum()}")

# Show unmatched variables that appear in many files (most important to document)
unmatched = compact[compact.ssb_beskrivelse.isna()].sort_values("n_files", ascending=False)
print(f"\nTop 20 unmatched variables (by frequency):")
for _, row in unmatched.head(20).iterrows():
    label = row.varlabel if pd.notna(row.varlabel) else "(no label)"
    print(f"  {row.varname:40s} in {row.n_files:3d} files  {label}")
