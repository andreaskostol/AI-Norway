"""
Build Felten et al. (2021) AIOE mapping to STYRK-08.

Three measures:
  - aioe: Overall AI Occupational Exposure (all 10 AI applications)
  - aioe_lm: Language Modeling AIOE (GenAI-relevant)
  - aioe_ig: Image Generation AIOE (GenAI-relevant)

Chain: SOC 2010 → ISCO-08 → STYRK-08

Data sources:
  - Felten et al. (2021, 2023) via https://github.com/AIOE-Data/AIOE
  - BLS SOC 2010 → ISCO-08 crosswalk
"""

import csv
from pathlib import Path
from collections import defaultdict

import openpyxl
import xlrd

BASE_DIR = Path(__file__).resolve().parents[3]
DATA_DIR = BASE_DIR / 'data' / 'ai_exposure'
FELTEN_DIR = DATA_DIR / 'felten'

AIOE_FILE = FELTEN_DIR / 'AIOE_DataAppendix.xlsx'
LM_FILE = FELTEN_DIR / 'Language_Modeling_AIOE_and_AIIE.xlsx'
IG_FILE = FELTEN_DIR / 'Image_Generation_AIOE_and_AIIE.xlsx'

SOC_ISCO_FILE = DATA_DIR / 'isco_soc_crosswalk.xls'
STYRK08_FILE = DATA_DIR / 'styrk08_codes.csv'

OUTPUT_FILE = DATA_DIR / 'styrk08_felten_mapping.csv'

MANUAL_STYRK_MAP = {
    '2223': '2221',  # Sykepleiere -> Nursing professionals
    '2224': '2221',  # Vernepleiere -> Nursing professionals
}

# Manual corrections for Norwegian STYRK adaptations where the same 4-digit
# BLS/ISCO code denotes a different occupation than the Norwegian STYRK code.
MANUAL_STYRK_SOC_MAP = {
    '2267': ['29-1122'],  # Ergoterapeuter -> Occupational Therapists
    '2269': ['29-1011'],  # Kiropraktorer mv. -> Chiropractors
}


def load_felten_aioe():
    """Load AIOE scores from Felten Data Appendix. Returns {soc: score}."""
    wb = openpyxl.load_workbook(str(AIOE_FILE))
    ws = wb['Appendix A']
    scores = {}
    for i in range(2, ws.max_row + 1):
        soc = str(ws.cell(i, 1).value).strip()
        val = ws.cell(i, 3).value
        if soc and '-' in soc and val is not None:
            scores[soc] = float(val)
    return scores


def load_felten_lm():
    """Load Language Modeling AIOE scores."""
    wb = openpyxl.load_workbook(str(LM_FILE))
    ws = wb['LM AIOE']
    scores = {}
    for i in range(2, ws.max_row + 1):
        soc = str(ws.cell(i, 1).value).strip()
        val = ws.cell(i, 3).value
        if soc and '-' in soc and val is not None:
            scores[soc] = float(val)
    return scores


def load_felten_ig():
    """Load Image Generation AIOE scores."""
    wb = openpyxl.load_workbook(str(IG_FILE))
    ws = wb['IG AIOE']
    scores = {}
    for i in range(2, ws.max_row + 1):
        soc = str(ws.cell(i, 1).value).strip()
        val = ws.cell(i, 3).value
        if soc and '-' in soc and val is not None:
            scores[soc] = float(val)
    return scores


def load_soc2010_to_isco08():
    """Load SOC 2010 -> ISCO-08 crosswalk with partial match flags."""
    wb = xlrd.open_workbook(str(SOC_ISCO_FILE))
    ws = wb.sheet_by_name('2010 SOC to ISCO-08')
    mapping: dict[str, list[str]] = defaultdict(list)
    partial: dict[tuple[str, str], bool] = {}
    for i in range(7, ws.nrows):
        soc = str(ws.cell_value(i, 0)).strip()
        part_flag = str(ws.cell_value(i, 2)).strip()
        isco = str(ws.cell_value(i, 3)).strip()
        if soc and isco and '-' in soc:
            if '.' in isco:
                isco = isco.split('.')[0]
            isco = isco.zfill(4)
            mapping[soc].append(isco)
            partial[(soc, isco)] = (part_flag == '*')
    return dict(mapping), partial


def load_styrk08_codes():
    codes = set()
    with open(STYRK08_FILE, encoding='latin-1') as f:
        for row in csv.DictReader(f):
            code = row.get('styrk08', row.get('code', ''))
            if len(code) == 4:
                codes.add(code)
    return codes


def crosswalk_soc_to_styrk(soc_scores, soc_to_isco, partial_flags, styrk_codes):
    """Crosswalk SOC 2010 scores to STYRK-08 with quality flags."""
    isco_contribs = defaultdict(list)
    for soc, score in soc_scores.items():
        if soc in soc_to_isco:
            fan_out = len(soc_to_isco[soc])
            for isco in soc_to_isco[soc]:
                is_partial = partial_flags.get((soc, isco), False)
                isco_contribs[isco].append({
                    'score': score,
                    'soc': soc,
                    'partial': is_partial,
                    'fan_out': fan_out if is_partial else 1,
                })

    results = {}
    for isco, contribs in isco_contribs.items():
        if isco not in styrk_codes:
            continue
        n = len(contribs)
        results[isco] = {
            'score': sum(c['score'] for c in contribs) / n,
            'n_soc_matched': n,
            'has_partial_match': 1 if any(c['partial'] for c in contribs) else 0,
            'max_partial_fanout': max((c['fan_out'] for c in contribs if c['partial']),
                                     default=0),
        }
    return results


def main():
    print("Loading Felten AIOE data...")
    aioe = load_felten_aioe()
    aioe_lm = load_felten_lm()
    aioe_ig = load_felten_ig()
    print(f"  AIOE: {len(aioe)} SOC codes")
    print(f"  LM AIOE: {len(aioe_lm)} SOC codes")
    print(f"  IG AIOE: {len(aioe_ig)} SOC codes")

    print("\nLoading crosswalks...")
    soc_to_isco, partial_flags = load_soc2010_to_isco08()
    styrk_codes = load_styrk08_codes()

    # Crosswalk each Felten measure
    print("\nCrosswalking AIOE...")
    aioe_styrk = crosswalk_soc_to_styrk(aioe, soc_to_isco, partial_flags, styrk_codes)
    print(f"  {len(aioe_styrk)} STYRK codes")

    print("Crosswalking LM AIOE...")
    lm_styrk = crosswalk_soc_to_styrk(aioe_lm, soc_to_isco, partial_flags, styrk_codes)
    print(f"  {len(lm_styrk)} STYRK codes")

    print("Crosswalking IG AIOE...")
    ig_styrk = crosswalk_soc_to_styrk(aioe_ig, soc_to_isco, partial_flags, styrk_codes)
    print(f"  {len(ig_styrk)} STYRK codes")

    # Merge all measures
    all_codes = set(aioe_styrk) | set(lm_styrk) | set(ig_styrk)
    results = []
    for code in sorted(all_codes):
        if code not in styrk_codes:
            continue
        r = {'styrk08': code}

        # AIOE measures (from SOC crosswalk)
        if code in aioe_styrk:
            a = aioe_styrk[code]
            r['aioe'] = a['score']
            r['n_soc_matched'] = a['n_soc_matched']
            r['has_partial_match'] = a['has_partial_match']
            r['max_partial_fanout'] = a['max_partial_fanout']
        else:
            r['aioe'] = ''
            r['n_soc_matched'] = 0
            r['has_partial_match'] = ''
            r['max_partial_fanout'] = ''

        r['aioe_lm'] = lm_styrk[code]['score'] if code in lm_styrk else ''
        r['aioe_ig'] = ig_styrk[code]['score'] if code in ig_styrk else ''

        r['manual_map'] = ''
        results.append(r)

    # Replace misleading same-code STYRK/ISCO matches with direct SOC sources.
    manual_targets = set(MANUAL_STYRK_SOC_MAP)
    results = [r for r in results if r['styrk08'] not in manual_targets]
    for styrk_target, soc_sources in MANUAL_STYRK_SOC_MAP.items():
        if styrk_target not in styrk_codes:
            continue
        used_socs = [soc for soc in soc_sources if soc in aioe]
        if not used_socs:
            print(f"  Manual SOC map skipped: {styrk_target} has no source scores")
            continue
        results.append({
            'styrk08': styrk_target,
            'aioe': sum(aioe[soc] for soc in used_socs) / len(used_socs),
            'aioe_lm': (sum(lm_s for soc in used_socs
                            if (lm_s := aioe_lm.get(soc)) is not None) /
                        max(1, sum(1 for soc in used_socs if soc in aioe_lm))),
            'aioe_ig': (sum(ig_s for soc in used_socs
                            if (ig_s := aioe_ig.get(soc)) is not None) /
                        max(1, sum(1 for soc in used_socs if soc in aioe_ig))),
            'n_soc_matched': len(used_socs),
            'has_partial_match': 0,
            'max_partial_fanout': 0,
            'manual_map': 'SOC:' + ';'.join(used_socs),
        })
        print(f"  Manual SOC map: {styrk_target} <- {';'.join(used_socs)}")

    # Manual mappings
    mapped_codes = {r['styrk08'] for r in results}
    for styrk_target, styrk_source in MANUAL_STYRK_MAP.items():
        if styrk_target in mapped_codes:
            continue
        if styrk_target not in styrk_codes:
            continue
        source_row = next((r for r in results if r['styrk08'] == styrk_source), None)
        if source_row:
            manual = dict(source_row)
            manual['styrk08'] = styrk_target
            manual['manual_map'] = styrk_source
            results.append(manual)
            print(f"  Manual map: {styrk_target} <- {styrk_source}")

    results.sort(key=lambda x: x['styrk08'])

    # Assign quintiles for each measure
    for measure in ['aioe', 'aioe_lm', 'aioe_ig']:
        codes_with_vals = [(r['styrk08'], r[measure]) for r in results
                          if r[measure] != '' and r[measure] is not None]
        if not codes_with_vals:
            continue
        codes_with_vals.sort(key=lambda x: x[1])
        n = len(codes_with_vals)
        quintiles = {}
        pctls = {}
        for i, (code, val) in enumerate(codes_with_vals):
            pctl = 100 * i / n
            pctls[code] = round(pctl, 2)
            if pctl <= 20: quintiles[code] = 1
            elif pctl <= 40: quintiles[code] = 2
            elif pctl <= 60: quintiles[code] = 3
            elif pctl <= 80: quintiles[code] = 4
            else: quintiles[code] = 5

        for r in results:
            r[f'pctl_{measure}'] = pctls.get(r['styrk08'], '')
            r[f'q_{measure}'] = quintiles.get(r['styrk08'], '')

    # Summary
    for measure in ['aioe', 'aioe_lm', 'aioe_ig']:
        n = sum(1 for r in results if r.get(measure, '') != '')
        vals = [r[measure] for r in results if r.get(measure, '') != '']
        if vals:
            from collections import Counter
            q_dist = Counter(r.get(f'q_{measure}', '') for r in results
                           if r.get(f'q_{measure}', '') != '')
            print(f"\n  {measure}: {n} codes, range {min(vals):.3f} to {max(vals):.3f}")
            print(f"    quintiles: {dict(sorted(q_dist.items()))}")

    n_partial = sum(1 for r in results if r.get('has_partial_match') == 1)
    print(f"\n  With partial SOC->ISCO match: {n_partial}")
    print(f"  Total STYRK codes: {len(results)}")

    # Save
    fieldnames = [
        'styrk08',
        'aioe', 'pctl_aioe', 'q_aioe',
        'aioe_lm', 'pctl_aioe_lm', 'q_aioe_lm',
        'aioe_ig', 'pctl_aioe_ig', 'q_aioe_ig',
        'n_soc_matched', 'has_partial_match', 'max_partial_fanout', 'manual_map',
    ]
    with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in results:
            out = {}
            for k in fieldnames:
                val = r.get(k, '')
                if isinstance(val, float):
                    out[k] = round(val, 6)
                else:
                    out[k] = val
            writer.writerow(out)

    print(f"\nSaved to {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
